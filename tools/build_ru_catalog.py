# -*- coding: utf-8 -*-
"""Build the Russian product catalog from the desktop product folder.

The source folder is organized as:
  level 1 category / level 2 category or series / level 3 product folder / files

Images whose names say they should be added to the description are processed as
description images and are not used as product card thumbnails.
"""

from __future__ import annotations

import html
import json
import os
import re
import shutil
from pathlib import Path
from urllib.parse import quote

from openpyxl import load_workbook
from PIL import Image, ImageOps


ROOT = Path.cwd()
SRC = Path(os.environ["PRODUCT_ROOT"])
IMG_OUT = ROOT / "assets" / "images" / "ru-products"
PAGE_OUT = ROOT / "ru" / "products"
SITE = "https://www.pratt-oil.com/"
DATE = "2026-06-16"
IMG_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".gif", ".tif", ".tiff"}
TABLE_EXTS = {".xlsx", ".xls"}
DESC_KEYS = [
    "添加到描述",
    "添加描述",
    "加到描述",
    "描述",
    "产品描述",
    "说明",
    "介绍",
    "注：",
    "注-",
    "注_",
    "описан",
    "description",
]
SKIP_KEYS = ["件号", "图号", "型号表"]

CYR = {
    "а": "a",
    "б": "b",
    "в": "v",
    "г": "g",
    "д": "d",
    "е": "e",
    "ё": "e",
    "ж": "zh",
    "з": "z",
    "и": "i",
    "й": "y",
    "к": "k",
    "л": "l",
    "м": "m",
    "н": "n",
    "о": "o",
    "п": "p",
    "р": "r",
    "с": "s",
    "т": "t",
    "у": "u",
    "ф": "f",
    "х": "h",
    "ц": "ts",
    "ч": "ch",
    "ш": "sh",
    "щ": "sch",
    "ъ": "",
    "ы": "y",
    "ь": "",
    "э": "e",
    "ю": "yu",
    "я": "ya",
}

CUSTOM_CATEGORIES = [
    "Буровые насосы и запасные части",
    "Буровые установки и принадлежности для лебедок",
    "Гидравлическое дисковое тормозное устройство серии PS",
    "WPT пневматическое фрикционное сцепление",
    "Устьевой инструмент",
    "Клапаны, насосы и гидравлические компоненты",
]


def reset_generated_dir(path: Path) -> None:
    resolved = path.resolve()
    if path.exists():
        if ROOT.resolve() not in resolved.parents:
            raise RuntimeError(f"Refusing to delete outside workspace: {resolved}")
        shutil.rmtree(path)
    path.mkdir(parents=True, exist_ok=True)


def clean_name(name: str) -> str:
    path = Path(name)
    known_suffixes = IMG_EXTS | TABLE_EXTS | {".docx", ".doc", ".pdf"}
    stem = path.stem if path.suffix.lower() in known_suffixes else name
    stem = re.sub(r"^\s*\d+\s*[\.．、_-]?\s*", "", stem)
    stem = re.sub(r"\s+", " ", stem).strip(" ._-")
    return stem or name


def natural_key(path: Path) -> tuple[int, str]:
    match = re.match(r"^\s*(\d+)", path.name)
    return (int(match.group(1)) if match else 9999, path.name.lower())


def slugify(text: str, fallback: str = "item") -> str:
    text = clean_name(text).lower()
    out: list[str] = []
    for ch in text:
        if ch in CYR:
            out.append(CYR[ch])
        elif ch.isascii() and ch.isalnum():
            out.append(ch)
        elif ch in ["+", "&"]:
            out.append(" and ")
        elif ch in [" ", "-", "_", "/", ".", ",", "(", ")"]:
            out.append("-")
    slug = "".join(out)
    slug = re.sub(r"[^a-z0-9-]+", "-", slug)
    slug = re.sub(r"-+", "-", slug).strip("-")
    return (slug[:96].strip("-") or fallback)


def unique_slug(base: str, used: set[str]) -> str:
    slug = base
    index = 2
    while slug in used:
        slug = f"{base}-{index}"
        index += 1
    used.add(slug)
    return slug


def is_desc_image(path: Path) -> bool:
    name = path.name.lower()
    return any(key.lower() in name for key in DESC_KEYS + SKIP_KEYS)


def safe_open_image(path: Path) -> Image.Image | None:
    try:
        image = Image.open(path)
        image = ImageOps.exif_transpose(image)
        return image.convert("RGBA")
    except Exception:
        return None


def image_score(path: Path) -> int:
    image = safe_open_image(path)
    if image is None:
        return -1
    width, height = image.size
    score = width * height
    name = path.name.lower()
    if is_desc_image(path):
        score -= 10**12
    if "微信" in name or "截图" in name:
        score -= 10**9
    if "清晰" in name or "处理" in name:
        score += 10**7
    if path.suffix.lower() in [".jpg", ".jpeg", ".png", ".webp"]:
        score += 10**6
    return score


def fit_image(src_path: Path, dest_path: Path, size: tuple[int, int], quality: int = 82) -> bool:
    image = safe_open_image(src_path)
    if image is None:
        return False
    width, height = image.size
    target_width, target_height = size
    ratio = min(target_width / width, target_height / height)
    new_size = (max(1, int(width * ratio)), max(1, int(height * ratio)))
    image = image.resize(new_size, Image.Resampling.LANCZOS)
    canvas = Image.new("RGBA", size, (248, 250, 249, 255))
    canvas.alpha_composite(image, ((target_width - new_size[0]) // 2, (target_height - new_size[1]) // 2))
    rgb = Image.new("RGB", size, (248, 250, 249))
    rgb.paste(canvas, mask=canvas.split()[-1])
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    rgb.save(dest_path, "WEBP", quality=quality, method=6)
    return True


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def extract_xlsx_table(path: Path, max_rows: int = 80, max_cols: int = 10) -> list[list[str]]:
    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return []
    try:
        for worksheet in workbook.worksheets:
            rows: list[list[str]] = []
            for row in worksheet.iter_rows(values_only=True):
                values = [clean_cell(value) for value in row[:max_cols]]
                while values and not values[-1]:
                    values.pop()
                if not values or not any(values):
                    continue
                rows.append(values)
                if len(rows) >= max_rows:
                    break
            if rows:
                return rows
    finally:
        workbook.close()
    return []


def extract_model_tables(product_dir: Path) -> tuple[list[dict], list[str]]:
    tables: list[dict] = []
    unsupported: list[str] = []
    table_files = sorted([path for path in product_dir.iterdir() if path.is_file() and path.suffix.lower() in TABLE_EXTS], key=natural_key)
    for table_file in table_files:
        if table_file.suffix.lower() == ".xlsx":
            rows = extract_xlsx_table(table_file)
            if rows:
                tables.append({"name": clean_name(table_file.name), "rows": rows})
            else:
                unsupported.append(table_file.name)
        else:
            unsupported.append(table_file.name)
    return tables, unsupported


def e(value: object) -> str:
    return html.escape(str(value), quote=True)


def whatsapp_link(product: str = "") -> str:
    message = "Здравствуйте Pratt Oil, мне нужна цена на нефтепромысловые запасные части"
    if product:
        message += f": {product}"
    return f"https://api.whatsapp.com/send?phone=8615908080040&text={quote(message)}"


def collect_products() -> tuple[list[dict], list[dict]]:
    reset_generated_dir(IMG_OUT)
    reset_generated_dir(PAGE_OUT)
    used_slugs: set[str] = set()
    products: list[dict] = []
    categories: list[dict] = []

    category_dirs = sorted([path for path in SRC.iterdir() if path.is_dir()], key=natural_key)
    for index, cat_dir in enumerate(category_dirs):
        raw_cat = clean_name(cat_dir.name)
        category_name = CUSTOM_CATEGORIES[index] if index < len(CUSTOM_CATEGORIES) else raw_cat
        category_slug = unique_slug(slugify(category_name, "category"), used_slugs)
        category = {"name": category_name, "raw": raw_cat, "slug": category_slug, "products": [], "children": {}}
        categories.append(category)

        product_dirs = [path for path in cat_dir.rglob("*") if path.is_dir()]
        product_dirs.sort(key=lambda path: [natural_key(part) for part in path.relative_to(cat_dir).parents][::-1] + [natural_key(path)])
        for product_dir in product_dirs:
            direct_images = [path for path in product_dir.iterdir() if path.is_file() and path.suffix.lower() in IMG_EXTS]
            direct_tables = [path for path in product_dir.iterdir() if path.is_file() and path.suffix.lower() in TABLE_EXTS]
            if not direct_images and not direct_tables:
                continue
            rel_parts = product_dir.relative_to(cat_dir).parts
            product_name = clean_name(product_dir.name)
            secondary = clean_name(rel_parts[0]) if len(rel_parts) > 1 else category_name
            slug = unique_slug(slugify(product_name, "product"), used_slugs)
            normal_images = [path for path in direct_images if not is_desc_image(path)]
            desc_images = [path for path in direct_images if is_desc_image(path)]
            model_tables, unsupported_tables = extract_model_tables(product_dir)

            main_image = None
            if normal_images:
                main = max(normal_images, key=image_score)
                out_name = f"{slug}-main.webp"
                if fit_image(main, IMG_OUT / out_name, (900, 650), quality=84):
                    main_image = f"assets/images/ru-products/{out_name}"

            description_images = []
            for desc_index, desc_path in enumerate(desc_images, 1):
                out_name = f"{slug}-description-{desc_index}.webp"
                if fit_image(desc_path, IMG_OUT / out_name, (1200, 800), quality=82):
                    description_images.append(f"assets/images/ru-products/{out_name}")

            product = {
                "name": product_name,
                "slug": slug,
                "category": category_name,
                "category_slug": category_slug,
                "secondary": secondary,
                "secondary_slug": slugify(secondary, "series"),
                "source_path": str(product_dir),
                "main_image": main_image,
                "description_images": description_images,
                "model_tables": model_tables,
                "unsupported_tables": unsupported_tables,
            }
            products.append(product)
            category["products"].append(product)
            category["children"].setdefault(secondary, []).append(product)

    return categories, products


def header(categories: list[dict], depth: int) -> str:
    if depth == 2:
        home = "../"
        products_href = "./"
        root_prefix = "../../"
    elif depth == 3:
        home = "../../"
        products_href = "../"
        root_prefix = "../../../"
    else:
        home = "./"
        products_href = "products/"
        root_prefix = "../"
    category_links = "".join(f'<a href="{products_href}#{e(cat["slug"])}">{e(cat["name"])}</a>' for cat in categories)
    return f"""<header class="site-header">
      <div class="topbar">
        <a href="{whatsapp_link()}" rel="nofollow">WhatsApp: +86 159 0808 0040</a>
        <a href="mailto:2000@pratt-oil.com">2000@pratt-oil.com</a>
        <a href="tel:+8615908080040">Phone: +86 159 0808 0040</a>
      </div>
      <nav class="nav" aria-label="Основная навигация">
        <a class="brand" href="{home}" aria-label="Jinan Prite Petroleum Equipment">
          <img class="brand-logo" src="{root_prefix}assets/images/logo/logo-pratt-oil.webp" alt="Pratt Oil logo" title="Pratt Oil logo" width="300" height="150" loading="eager" decoding="async">
          <span><strong>PRATT OIL</strong><small>Jinan Prite Petroleum Equipment</small></span>
        </a>
        <button class="menu-toggle" type="button" aria-label="Open navigation" aria-expanded="false">Меню</button>
        <div class="nav-links" id="navLinks">
          <a class="nav-main-link" href="{home}">Главная</a>
          <div class="nav-dropdown" data-nav-dropdown>
            <div class="nav-dropdown-row"><a class="nav-main-link" href="{root_prefix}en/about-us/" aria-haspopup="true">О нас</a><button class="nav-dropdown-toggle" type="button" data-nav-toggle aria-label="Открыть меню о компании" aria-expanded="false">▾</button></div>
            <div class="dropdown-panel" data-nav-panel><a href="{root_prefix}en/about-us/">Профиль компании</a><a href="{root_prefix}en/factory-photos/">Фотографии завода</a><a href="{root_prefix}en/export-support/">Экспортная поддержка</a></div>
          </div>
          <div class="nav-dropdown" data-nav-dropdown>
            <div class="nav-dropdown-row"><a class="nav-main-link" href="{products_href}" aria-haspopup="true">Продукция</a><button class="nav-dropdown-toggle" type="button" data-nav-toggle aria-label="Открыть меню продукции" aria-expanded="false">▾</button></div>
            <div class="dropdown-panel" data-nav-panel>{category_links}</div>
          </div>
          <div class="nav-dropdown" data-nav-dropdown>
            <div class="nav-dropdown-row"><a class="nav-main-link" href="{root_prefix}en/news/" aria-haspopup="true">Новости</a><button class="nav-dropdown-toggle" type="button" data-nav-toggle aria-label="Открыть меню новостей" aria-expanded="false">▾</button></div>
            <div class="dropdown-panel" data-nav-panel><a href="{root_prefix}en/news/company-news/">Новости компании</a><a href="{root_prefix}en/news/industry-news/">Новости отрасли</a></div>
          </div>
          <a class="nav-main-link" href="{root_prefix}en/services/">Услуги</a>
          <a class="nav-main-link" href="{root_prefix}en/contact/">Контакты</a>
        </div>
        <div class="nav-actions">
          <div class="language-dropdown" data-nav-dropdown>
            <button class="language-button" type="button" data-nav-toggle aria-label="Open language menu" aria-expanded="false">Language</button>
            <div class="dropdown-panel language-panel" data-nav-panel><a class="language-option" href="{root_prefix}">English</a><a class="language-option" href="{root_prefix}ar/">Arabic</a><a class="language-option active" href="{root_prefix}ru/">Russian</a></div>
          </div>
          <a class="nav-quick-link" href="{whatsapp_link()}" rel="nofollow">WhatsApp</a>
        </div>
      </nav>
    </header>"""


def footer(depth: int) -> str:
    root_prefix = "../../" if depth == 2 else "../../../"
    ru_prefix = "../" if depth == 2 else "../../"
    return f"""<footer class="footer">
      <div class="footer-brand"><img class="footer-logo" src="{root_prefix}assets/images/logo/logo-pratt-oil.webp" alt="Pratt Oil logo" title="Pratt Oil logo" width="300" height="150" loading="lazy" decoding="async"><div><strong>Jinan Prite Petroleum Equipment Co., Ltd.</strong><p>Поставщик нефтебурового оборудования и запасных частей для Казахстана, Узбекистана, Туркменистана, Кыргызстана, Таджикистана и других рынков Центральной Азии.</p></div></div>
      <div class="footer-links"><a href="{ru_prefix}#products">Продукция</a><a href="{ru_prefix}#company-presence">Возможности</a><a href="{ru_prefix}#contact">Контакты</a><a href="{whatsapp_link()}" rel="nofollow">WhatsApp</a></div>
    </footer>
    <a class="floating-whatsapp" href="{whatsapp_link()}" rel="nofollow" aria-label="WhatsApp inquiry">WhatsApp</a>"""


def page_head(title: str, description: str, canonical: str, image: str | None, depth: int, schema: dict | None) -> str:
    root_prefix = "../../" if depth == 2 else "../../../"
    image_path = image or "assets/images/hero/hero-oilfield-equipment-spare-parts-middle-east.webp"
    full_image = SITE + image_path
    schema_tag = ""
    if schema:
        schema_tag = f'\n    <script type="application/ld+json">{json.dumps(schema, ensure_ascii=False, separators=(",", ":"))}</script>'
    return f"""<!doctype html>
<html lang="ru">
  <head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>{e(title)}</title>
    <meta name="description" content="{e(description)}">
    <meta name="robots" content="index, follow, max-image-preview:large">
    <link rel="canonical" href="{e(canonical)}">
<link rel="alternate" hreflang="x-default" href="{e(SITE)}ru/products/">
    <meta property="og:type" content="website">
    <meta property="og:title" content="{e(title)}">
    <meta property="og:description" content="{e(description)}">
    <meta property="og:url" content="{e(canonical)}">
    <meta property="og:site_name" content="Jinan Prite Petroleum Equipment Co., Ltd.">
    <meta property="og:locale" content="ru_RU">
    <meta property="og:image" content="{e(full_image)}">
    <meta name="twitter:card" content="summary_large_image">
    <meta name="twitter:title" content="{e(title)}">
    <meta name="twitter:description" content="{e(description)}">
    <meta name="twitter:image" content="{e(full_image)}">
    <link rel="stylesheet" href="{root_prefix}styles.css?v=ru-products-20260616">{schema_tag}
  </head>"""


def build_catalog_page(categories: list[dict], products: list[dict]) -> None:
    category_cards = []
    for category in categories:
        thumbnail = next((product["main_image"] for product in category["products"] if product["main_image"]), None)
        image_html = (
            f'<img src="../../{e(thumbnail)}" alt="{e(category["name"])}" title="{e(category["name"])}" width="900" height="650" loading="lazy" decoding="async">'
            if thumbnail
            else "<span>Фото категории</span>"
        )
        category_cards.append(
            f'<a class="ru-category-card" href="#{e(category["slug"])}"><div class="ru-category-card-media">{image_html}</div><strong>{e(category["name"])}</strong><span>{len(category["products"])} товаров</span></a>'
        )

    sections = []
    for category in categories:
        subcategory_links = "".join(
            f'<a href="#{e(category["slug"])}-{e(slugify(secondary, "series"))}">{e(secondary)}</a>'
            for secondary in category["children"].keys()
        )
        blocks = []
        for secondary, group_products in category["children"].items():
            cards = []
            for product in group_products:
                image_html = (
                    f'<img src="../../{e(product["main_image"])}" alt="{e(product["name"])}" title="{e(product["name"])}" width="900" height="650" loading="lazy" decoding="async">'
                    if product["main_image"]
                    else "<span>Фото товара будет загружено</span>"
                )
                cards.append(
                    f"""<article class="ru-product-card">
              <a href="{e(product["slug"])}/">
                <div class="ru-product-card-media">{image_html}</div>
                <div class="ru-product-card-body"><span>{e(secondary)}</span><h3>{e(product["name"])}</h3><p>{e(category["name"])}</p></div>
              </a>
              <a class="mini-btn dark" href="{whatsapp_link(product["name"])}" rel="nofollow">Запросить цену</a>
            </article>"""
                )
            blocks.append(
                f"""<div class="ru-subcategory-block" id="{e(category["slug"])}-{e(slugify(secondary, "series"))}">
            <h3>{e(secondary)}</h3>
            <div class="ru-product-grid">{''.join(cards)}</div>
          </div>"""
            )
        sections.append(
            f"""<section class="ru-category-section" id="{e(category["slug"])}">
          <div class="ru-category-heading"><p class="eyebrow">Категория продукции</p><h2>{e(category["name"])}</h2><p>Запасные части и оборудование для буровых проектов Казахстана и стран Центральной Азии. Отправьте номер детали, модель или фото для подбора.</p></div>
          <div class="ru-subcategory-links">{subcategory_links}</div>
          {''.join(blocks)}
        </section>"""
        )

    schema = {
        "@context": "https://schema.org",
        "@graph": [
            {
                "@type": "Organization",
                "@id": SITE + "#organization",
                "name": "Jinan Prite Petroleum Equipment Co., Ltd.",
                "alternateName": "Pratt Oil",
                "url": SITE,
                "email": ["2000@pratt-oil.com", "2001@pratt-oil.com"],
                "telephone": ["+86-15908080040", "+86-18063410221"],
            },
            {
                "@type": "CollectionPage",
                "@id": SITE + "ru/products/#webpage",
                "url": SITE + "ru/products/",
                "name": "Каталог нефтепромыслового оборудования и запасных частей",
                "inLanguage": "ru",
            },
            {
                "@type": "BreadcrumbList",
                "itemListElement": [
                    {"@type": "ListItem", "position": 1, "name": "Главная", "item": SITE + "ru/"},
                    {"@type": "ListItem", "position": 2, "name": "Продукция", "item": SITE + "ru/products/"},
                ],
            },
        ],
    }
    html_text = (
        page_head(
            "Каталог нефтепромыслового оборудования и запасных частей | Pratt Oil",
            "Русский каталог запасных частей для буровых насосов, буровых установок, тормозных систем, WPT сцеплений, устьевого инструмента и гидравлических компонентов для Казахстана и Центральной Азии.",
            SITE + "ru/products/",
            next((product["main_image"] for product in products if product["main_image"]), None),
            2,
            schema,
        )
        + f"""
  <body class="ru-catalog-page">
    {header(categories, 2)}
    <main>
      <section class="ru-catalog-hero">
        <p class="eyebrow">Русский каталог продукции</p>
        <h1>Нефтепромысловое оборудование и запасные части для Казахстана и Центральной Азии</h1>
        <p>Каталог сформирован по папкам поставщика: первый уровень, второй уровень и конкретные товары сохранены как структура продукции. Названия продуктов взяты из имен файлов и папок.</p>
        <div class="hero-actions"><a class="btn primary" href="{whatsapp_link()}" rel="nofollow">Запросить цену в WhatsApp</a><a class="btn secondary light" href="#catalog">Смотреть продукцию</a></div>
      </section>
      <section class="section ru-category-overview" id="catalog">
        <div class="section-heading"><p class="eyebrow">Продукция</p><h2>Основные категории</h2><p>Выберите категорию или прокрутите каталог ниже. Все изображения обработаны в едином размере WebP.</p></div>
        <div class="ru-category-card-grid">{''.join(category_cards)}</div>
      </section>
      <section class="section ru-catalog-sections">
        {''.join(sections)}
      </section>
      <section class="section contact-section" id="contact">
        <div><p class="eyebrow">Запрос</p><h2>Отправьте номер детали или фото продукта</h2><p>Для быстрого подбора отправьте номер детали, модель насоса, модель буровой установки, фото продукта или технический чертеж.</p></div>
        <div class="contact-grid"><a href="mailto:2000@pratt-oil.com"><span>Email</span><strong>2000@pratt-oil.com</strong></a><a href="mailto:2001@pratt-oil.com"><span>Email</span><strong>2001@pratt-oil.com</strong></a><a href="{whatsapp_link()}" rel="nofollow"><span>WhatsApp</span><strong>+86 159 0808 0040</strong></a><a href="tel:+8618063410221"><span>Phone</span><strong>+86 180 6341 0221</strong></a></div>
        <address>Building 11, Zhigu Industrial Park, High-tech Zone, Jinan, Shandong Province, China</address>
      </section>
    </main>
    {footer(2)}
    <script src="../../script.js"></script>
  </body>
</html>
"""
    )
    (PAGE_OUT / "index.html").write_text(html_text, encoding="utf-8")


def render_model_tables(product: dict) -> str:
    tables = []
    for table in product.get("model_tables", []):
        rows = table.get("rows", [])
        if not rows:
            continue
        max_cols = max(len(row) for row in rows)
        normalized = [row + [""] * (max_cols - len(row)) for row in rows]
        head = "".join(f"<th>{e(value)}</th>" for value in normalized[0])
        body_rows = []
        for row in normalized[1:]:
            body_rows.append("<tr>" + "".join(f"<td>{e(value)}</td>" for value in row) + "</tr>")
        if not body_rows and normalized:
            body_rows.append("<tr>" + "".join(f"<td>{e(value)}</td>" for value in normalized[0]) + "</tr>")
            head = "".join(f"<th>Колонка {index}</th>" for index in range(1, max_cols + 1))
        tables.append(
            f"""<div class="ru-model-table-wrap">
          <h3>{e(table.get("name", "Таблица моделей"))}</h3>
          <div class="ru-table-scroll"><table class="ru-model-table"><thead><tr>{head}</tr></thead><tbody>{''.join(body_rows)}</tbody></table></div>
        </div>"""
        )
    unsupported = product.get("unsupported_tables", [])
    if unsupported:
        files = "".join(f"<li>{e(name)}</li>" for name in unsupported)
        tables.append(
            f"""<div class="ru-model-table-wrap ru-unsupported-table-note">
          <h3>Исходные файлы таблиц</h3>
          <p>Эти файлы есть в папке товара, но сохранены в старом формате или требуют ручной проверки перед публикацией:</p>
          <ul>{files}</ul>
        </div>"""
        )
    if not tables:
        return '<p class="ru-muted">Таблица моделей пока не загружена. Отправьте номер детали, модель оборудования или фото для подбора.</p>'
    return "\n        ".join(tables)


def build_detail_pages(categories: list[dict], products: list[dict]) -> None:
    for product in products:
        product_dir = PAGE_OUT / product["slug"]
        product_dir.mkdir(parents=True, exist_ok=True)
        image_html = (
            f'<img src="../../../{e(product["main_image"])}" alt="{e(product["name"])}" title="{e(product["name"])}" width="900" height="650" loading="eager" decoding="async">'
            if product["main_image"]
            else "<span>Фото товара будет загружено</span>"
        )
        description_images = "".join(
            f'<figure><img src="../../../{e(desc)}" alt="{e(product["name"])} описание {index}" title="{e(product["name"])} описание {index}" width="1200" height="800" loading="lazy" decoding="async"><figcaption>{e(product["name"])} - описание {index}</figcaption></figure>'
            for index, desc in enumerate(product["description_images"], 1)
        )
        if not description_images:
            description_images = '<p class="ru-muted">Дополнительные фотографии описания пока не загружены. Отправьте номер детали или фото для уточнения спецификации.</p>'
        model_tables = render_model_tables(product)
        related = [item for item in products if item["category_slug"] == product["category_slug"] and item["slug"] != product["slug"]][:4]
        related_html = "".join(f'<a href="../{e(item["slug"])}/">{e(item["name"])}</a>' for item in related)
        image_url = product["main_image"] or "assets/images/hero/hero-oilfield-equipment-spare-parts-middle-east.webp"
        schema = {
            "@context": "https://schema.org",
            "@graph": [
                {"@type": "Organization", "@id": SITE + "#organization", "name": "Jinan Prite Petroleum Equipment Co., Ltd.", "alternateName": "Pratt Oil", "url": SITE},
                {
                    "@type": "Product",
                    "name": product["name"],
                    "category": product["category"],
                    "brand": {"@type": "Brand", "name": "Pratt Oil"},
                    "image": SITE + image_url,
                    "description": f'{product["name"]} - {product["category"]} для нефтепромыслового оборудования. Запросите цену по номеру детали, модели или фото продукта.',
                },
                {
                    "@type": "BreadcrumbList",
                    "itemListElement": [
                        {"@type": "ListItem", "position": 1, "name": "Главная", "item": SITE + "ru/"},
                        {"@type": "ListItem", "position": 2, "name": "Продукция", "item": SITE + "ru/products/"},
                        {"@type": "ListItem", "position": 3, "name": product["category"], "item": SITE + "ru/products/#" + product["category_slug"]},
                        {"@type": "ListItem", "position": 4, "name": product["name"], "item": SITE + "ru/products/" + product["slug"] + "/"},
                    ],
                },
            ],
        }
        html_text = (
            page_head(
                f'{product["name"]} | Pratt Oil',
                f'{product["name"]}: поставка и подбор запасных частей для нефтепромыслового оборудования. Отправьте номер детали, модель или фото для расчета цены.',
                SITE + f'ru/products/{product["slug"]}/',
                image_url,
                3,
                schema,
            )
            + f"""
  <body class="ru-product-detail-page">
    {header(categories, 3)}
    <main>
      <nav class="ru-breadcrumb" aria-label="Breadcrumb"><a href="../../">Главная</a><span>›</span><a href="../">Продукция</a><span>›</span><a href="../#{e(product["category_slug"])}">{e(product["category"])}</a><span>›</span><strong>{e(product["name"])}</strong></nav>
      <section class="section ru-product-detail-hero">
        <div class="ru-product-detail-copy">
          <p class="eyebrow">{e(product["category"])}</p>
          <h1>{e(product["name"])}</h1>
          <p>{e(product["secondary"])}. Поставка для буровых, ремонтных и нефтесервисных проектов в Казахстане и странах Центральной Азии.</p>
          <div class="ru-product-meta"><span>Первый уровень: {e(product["category"])}</span><span>Второй уровень: {e(product["secondary"])}</span><span>Товар: {e(product["name"])}</span></div>
          <div class="hero-actions"><a class="btn primary" href="{whatsapp_link(product["name"])}" rel="nofollow">Запросить цену в WhatsApp</a><a class="btn secondary dark-text" href="mailto:2000@pratt-oil.com">Отправить Email</a></div>
        </div>
        <div class="ru-product-detail-image">{image_html}</div>
      </section>
      <section class="section ru-product-info-grid">
        <article><h2>Как запросить цену</h2><ul><li>Отправьте номер детали или модель оборудования.</li><li>Можно приложить фото продукта или технический чертеж.</li><li>Мы проверим совместимость и подготовим предложение.</li></ul></article>
        <article><h2>Применение</h2><p>Используется для нефтепромыслового оборудования, буровых установок, насосных систем, устьевых операций, тормозных и гидравлических узлов в зависимости от категории продукта.</p></article>
        <article><h2>Поддержка поставки</h2><p>Подбор спецификации, экспортная упаковка, коммуникация на русском языке, WhatsApp и email поддержка для покупателей Центральной Азии.</p></article>
      </section>
      <section class="section ru-description-images">
        <div class="section-heading"><p class="eyebrow">Описание</p><h2>Фотографии и материалы описания</h2><p>Изображения с пометкой «добавить в описание» размещены здесь, а не используются как главная фотография товара.</p></div>
        <div class="ru-description-grid">{description_images}</div>
      </section>
      <section class="section ru-model-table-section">
        <div class="section-heading"><p class="eyebrow">Модели и спецификации</p><h2>Таблица моделей</h2><p>Если в исходной папке есть таблица с моделями, она отображается здесь для удобства подбора и запроса цены.</p></div>
        {model_tables}
      </section>
      <section class="section ru-related-products"><div class="section-heading"><p class="eyebrow">Похожие товары</p><h2>Другие товары в этой категории</h2></div><div class="ru-related-links">{related_html}</div></section>
      <section class="section contact-section" id="contact"><div><p class="eyebrow">Запрос</p><h2>Связаться с Pratt Oil</h2><p>Отправьте номер детали, модель, фото продукта или технический чертеж. Мы проверим совместимость и ответим с деталями предложения.</p></div><div class="contact-grid"><a href="mailto:2000@pratt-oil.com"><span>Email</span><strong>2000@pratt-oil.com</strong></a><a href="mailto:2001@pratt-oil.com"><span>Email</span><strong>2001@pratt-oil.com</strong></a><a href="{whatsapp_link(product["name"])}" rel="nofollow"><span>WhatsApp</span><strong>+86 159 0808 0040</strong></a><div><span>WeChat</span><strong>15908080040</strong></div></div><address>Building 11, Zhigu Industrial Park, High-tech Zone, Jinan, Shandong Province, China</address></section>
    </main>
    {footer(3)}
    <script src="../../../script.js"></script>
  </body>
</html>
"""
        )
        (product_dir / "index.html").write_text(html_text, encoding="utf-8")


def update_ru_home(categories: list[dict]) -> None:
    cards = []
    for category in categories:
        thumbnail = next((product["main_image"] for product in category["products"] if product["main_image"]), None)
        image_html = (
            f'<img class="product-card-image" src="../{e(thumbnail)}" alt="{e(category["name"])}" title="{e(category["name"])}" width="900" height="650" loading="lazy" decoding="async">'
            if thumbnail
            else "<span>Фото категории</span>"
        )
        cards.append(
            f"""          <article class="product-card">
            <a class="card-detail-link" href="products/#{e(category["slug"])}">
              <div class="product-card-media has-image">{image_html}</div>
              <h3>{e(category["name"])}</h3>
              <p>Категория включает {len(category["products"])} товаров и запасных частей. Отправьте номер детали, модель или фото для подбора.</p>
            </a>
            <div class="card-actions"><a class="mini-btn" href="products/#{e(category["slug"])}">Подробнее</a><a class="mini-btn" href="#contact">Запросить цену</a><a class="mini-btn dark" href="{whatsapp_link(category["name"])}" rel="nofollow">WhatsApp</a></div>
          </article>"""
        )
    home_cards = "\n".join(cards)
    index_path = ROOT / "ru" / "index.html"
    text = index_path.read_text(encoding="utf-8")
    text = text.replace('<link rel="stylesheet" href="../styles.css?v=ar-ru-mobile-ui-20260615">', '<link rel="stylesheet" href="../styles.css?v=ru-products-20260616">')
    text = text.replace('<link rel="stylesheet" href="../styles.css?v=ru-products-20260616">', '<link rel="stylesheet" href="../styles.css?v=ru-products-20260616">')
    text = text.replace("<body>", '<body class="home-ru">', 1)
    text = text.replace('href="../index.html#products" aria-haspopup="true">Продукция</a>', 'href="products/" aria-haspopup="true">Продукция</a>')
    text = text.replace('href="products/" aria-haspopup="true">Продукция</a>', 'href="products/" aria-haspopup="true">Продукция</a>')
    category_links = '<div class="dropdown-panel" data-nav-panel>' + "".join(f'<a href="products/#{e(category["slug"])}">{e(category["name"])}</a>' for category in categories) + "</div>"
    pattern = re.compile(
        r'(<div class="nav-dropdown" data-nav-dropdown>\s*<div class="nav-dropdown-row">\s*<a class="nav-main-link" href="products/" aria-haspopup="true">Продукция</a>.*?</div>\s*)<div class="dropdown-panel" data-nav-panel>.*?</div>',
        flags=re.S,
    )
    text = pattern.sub(r"\1" + category_links, text, count=1)
    start = text.index('      <section class="section product-entry-section" id="products">')
    end = text.index('      <section class="section company-presence-section" id="company-presence">', start)
    new_section = f"""      <section class="section product-entry-section" id="products">
        <div class="section-heading">
          <p class="eyebrow">Наша продукция</p>
          <h2>Категории нефтепромыслового оборудования и запасных частей</h2>
          <p>Каталог для покупателей Казахстана и стран Центральной Азии: буровые насосы, буровые установки, тормозные системы, WPT сцепления, устьевой инструмент, клапаны, насосы и гидравлические компоненты.</p>
        </div>
        <div class="product-grid">
{home_cards}
        </div>
        <div class="product-more-row"><a class="btn product-more-btn" href="products/">Смотреть весь каталог</a></div>
      </section>
"""
    text = text[:start] + new_section + text[end:]
    index_path.write_text(text, encoding="utf-8")


def update_sitemap(products: list[dict]) -> None:
    sitemap = ROOT / "sitemap.xml"
    text = sitemap.read_text(encoding="utf-8")
    text = re.sub(r"\s*<url>\s*<loc>https://www\.pratt-oil\.com/ru/products/.*?</url>", "", text, flags=re.S)
    urls = [SITE + "ru/products/"] + [SITE + f'ru/products/{product["slug"]}/' for product in products]
    entries = []
    for url in urls:
        priority = "0.8" if url.endswith("/products/") else "0.6"
        entries.append(f"  <url>\n    <loc>{url}</loc>\n    <lastmod>{DATE}</lastmod>\n    <changefreq>weekly</changefreq>\n    <priority>{priority}</priority>\n  </url>")
    text = text.replace("</urlset>", "\n" + "\n".join(entries) + "\n</urlset>")
    sitemap.write_text(text, encoding="utf-8")


def write_manifest(categories: list[dict], products: list[dict]) -> None:
    (ROOT / "work").mkdir(exist_ok=True)
    manifest = {
        "source": str(SRC),
        "categories": [{"name": item["name"], "slug": item["slug"], "product_count": len(item["products"])} for item in categories],
        "product_count": len(products),
        "main_images": sum(1 for item in products if item["main_image"]),
        "description_images": sum(len(item["description_images"]) for item in products),
        "model_tables": sum(len(item["model_tables"]) for item in products),
        "unsupported_table_files": sum(len(item["unsupported_tables"]) for item in products),
        "products_without_main_image": [item["name"] for item in products if not item["main_image"]],
    }
    (ROOT / "work" / "ru_products_manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(manifest, ensure_ascii=False, indent=2))


def main() -> None:
    categories, products = collect_products()
    build_catalog_page(categories, products)
    build_detail_pages(categories, products)
    update_ru_home(categories)
    update_sitemap(products)
    write_manifest(categories, products)


if __name__ == "__main__":
    main()
